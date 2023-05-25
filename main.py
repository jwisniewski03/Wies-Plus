import pandas as pd
import openai
from openpyxl.reader.excel import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.workbook import Workbook
from config import AUTH_KEY


# Wczytanie danych z pliku
def wczytaj_dane(file):
    df = pd.read_excel(file)
    return df


# Zapis danych do pliku
def zapisz_dane(output_file, nazwa, gmina, powiat, opis):
    df = pd.DataFrame(zip(nazwa.split(), gmina.split(), powiat.split(), opis.split(None, 0)),
                      columns=["Nazwa", "Gmina", "Powiat", "Opis"])

    try:
        wb = load_workbook(filename=output_file)
    except FileNotFoundError:
        print("Error - file not found... ")
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        print("New file created.")
        ws.append(["Nazwa", "Gmina", "Powiat", "Opis"])  # Column names
    else:
        ws = wb["Sheet1"]

    for r in dataframe_to_rows(df, index=False, header=False):
        ws.append(r)

    wb.save(output_file)
    return None


# Generowanie opisu wsi za pomocą GPT-4
def generuj_opis_wsi(nazwa, gmina, powiat):
    try:
        # Make your OpenAI API request here
        # Inicjalizacja GPT-4
        openai.api_key = AUTH_KEY
        res = openai.ChatCompletion.create(
            model="gpt-3.5-turbo",
            messages=[
                {"role": "user",
                 "content": "Opisz wieś " + nazwa + " znajdującej się w gminie " + gmina + " w powiecie :" + powiat}
            ]
        )
        return res["choices"][0]["message"]["content"]

    except openai.error.APIError as e:
        # Handle API error here, e.g. retry or log
        print(f"OpenAI API returned an API Error: {e}")
        pass
    except openai.error.APIConnectionError as e:
        # Handle connection error here
        print(f"Failed to connect to OpenAI API: {e}")
        pass
    except openai.error.RateLimitError as e:
        # Handle rate limit error (we recommend using exponential backoff)
        print(f"OpenAI API request exceeded rate limit: {e}")
        pass


# # Przykładowe dane
input_file = 'dane_wsi.xlsx'
dane_wsi = wczytaj_dane(input_file)

for _, wiersz in dane_wsi.iterrows():
    nazwa = wiersz['Nazwa']
    gmina = wiersz['Gmina']
    powiat = wiersz['Powiat']

    wyniki = "".join(generuj_opis_wsi(nazwa, gmina, powiat).split())

    #Zapis danych wyjściowych do wybranego pliku
    output_file = 'wyniki.xlsx'
    zapisz_dane(output_file, nazwa, gmina, powiat, wyniki)
